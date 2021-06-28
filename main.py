from openpyxl.styles import Alignment, Font, PatternFill

from pyfire import ModelRow, RowStyle, Section, Style, Total, Model

WHITE = "00FFFFFF"
# Income
DARK_GREEN = "00375623"
LIGHT_GREEN = "00E2EFDA"
MEDIUM_GREEN = "00A9D08E"
TEAL = "0064E7E3"
# Expenses
DARK_GRAY = "00525252"
LIGHT_GRAY = "00D0CECE"
MEDIUM_GRAY = "00AEAAAA"
LIGHT_PURPLE = "009791EB"
# Assets
DARK_ORANGE = "00833C0C"
LIGHT_ORANGE = "00FCE4D6"
MEDIUM_ORANGE = "00F4B084"
# Liabilities
DARK_BLUE = "001F4E78"
LIGHT_BLUE = "00DDEBF7"
MEDIUM_BLUE = "009BC2E6"

INCOME_HEADER_STYLE = RowStyle(
    label_style=Style(
        font=Font(color=WHITE, bold=True),
        alignment=Alignment(horizontal="center"),
        fill=PatternFill(patternType="solid", fgColor=DARK_GREEN),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=DARK_GREEN)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=DARK_GREEN)),
)
INCOME_BODY_STYLE = RowStyle(
    label_style=Style(
        alignment=Alignment(indent=1),
        fill=PatternFill(patternType="solid", fgColor=LIGHT_GREEN),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GREEN)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GREEN)),
)
INCOME_SUBTOTAL_STYLE = RowStyle(
    label_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GREEN)),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GREEN)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GREEN)),
)
INCOME_TOTAL_STYLE = RowStyle(
    label_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_GREEN)),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_GREEN)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_GREEN)),
)
INCOME_ADJUSTED_TOTAL_STYLE = RowStyle(
    label_style=Style(
        fill=PatternFill(patternType="solid", fgColor=TEAL),
        alignment=Alignment(horizontal="center"),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=TEAL)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=TEAL)),
)

EXPENSE_HEADER_STYLE = RowStyle(
    label_style=Style(
        font=Font(color=WHITE, bold=True),
        alignment=Alignment(horizontal="center"),
        fill=PatternFill(patternType="solid", fgColor=DARK_GRAY),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=DARK_GRAY)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=DARK_GRAY)),
)
EXPENSE_BODY_STYLE = RowStyle(
    label_style=Style(
        alignment=Alignment(indent=1),
        fill=PatternFill(patternType="solid", fgColor=LIGHT_GRAY),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GRAY)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GRAY)),
)
EXPENSE_SUBTOTAL_STYLE = RowStyle(
    label_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GRAY)),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GRAY)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_GRAY)),
)
EXPENSE_TOTAL_STYLE = RowStyle(
    label_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_GRAY)),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_GRAY)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_GRAY)),
)
EXPENSE_ADJUSTED_TOTAL_STYLE = RowStyle(
    label_style=Style(
        fill=PatternFill(patternType="solid", fgColor=LIGHT_PURPLE),
        alignment=Alignment(horizontal="center"),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_PURPLE)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_PURPLE)),
)

ASSET_HEADER_STYLE = RowStyle(
    label_style=Style(
        font=Font(color=WHITE, bold=True),
        alignment=Alignment(horizontal="center"),
        fill=PatternFill(patternType="solid", fgColor=DARK_ORANGE),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=DARK_ORANGE)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=DARK_ORANGE)),
)
ASSET_BODY_STYLE = RowStyle(
    label_style=Style(
        alignment=Alignment(indent=1),
        fill=PatternFill(patternType="solid", fgColor=LIGHT_ORANGE),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_ORANGE)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_ORANGE)),
)
ASSET_SUBTOTAL_STYLE = RowStyle(
    label_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_ORANGE)),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_ORANGE)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_ORANGE)),
)
ASSET_TOTAL_STYLE = RowStyle(
    label_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_ORANGE)),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_ORANGE)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_ORANGE)),
)

LIABILITY_HEADER_STYLE = RowStyle(
    label_style=Style(
        font=Font(color=WHITE, bold=True),
        alignment=Alignment(horizontal="center"),
        fill=PatternFill(patternType="solid", fgColor=DARK_BLUE),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=DARK_BLUE)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=DARK_BLUE)),
)
LIABILITY_BODY_STYLE = RowStyle(
    label_style=Style(
        alignment=Alignment(indent=1),
        fill=PatternFill(patternType="solid", fgColor=LIGHT_BLUE),
    ),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_BLUE)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_BLUE)),
)
LIABILITY_SUBTOTAL_STYLE = RowStyle(
    label_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_BLUE)),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_BLUE)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=LIGHT_BLUE)),
)
LIABILITY_TOTAL_STYLE = RowStyle(
    label_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_BLUE)),
    data_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_BLUE)),
    row_total_style=Style(fill=PatternFill(patternType="solid", fgColor=MEDIUM_BLUE)),
)


def build_model_tree(data: dict) -> ModelRow:
    _section_style_defaults = {
        "Income": {
            "Header": INCOME_HEADER_STYLE,
            "Body": INCOME_BODY_STYLE,
            "Subtotal": INCOME_SUBTOTAL_STYLE,
            "Total": INCOME_TOTAL_STYLE,
            "Adjusted Total": INCOME_ADJUSTED_TOTAL_STYLE,
        },
        "Expenses": {
            "Header": EXPENSE_HEADER_STYLE,
            "Body": EXPENSE_BODY_STYLE,
            "Subtotal": EXPENSE_SUBTOTAL_STYLE,
            "Total": EXPENSE_TOTAL_STYLE,
            "Adjusted Total": EXPENSE_ADJUSTED_TOTAL_STYLE,
        },
        "Assets": {
            "Header": ASSET_HEADER_STYLE,
            "Body": ASSET_BODY_STYLE,
            "Subtotal": ASSET_SUBTOTAL_STYLE,
            "Total": ASSET_TOTAL_STYLE,
        },
        "Liabilities": {
            "Header": LIABILITY_HEADER_STYLE,
            "Body": LIABILITY_BODY_STYLE,
            "Subtotal": LIABILITY_SUBTOTAL_STYLE,
            "Total": LIABILITY_TOTAL_STYLE,
        },
    }
    _section_kwarg_defaults = {
        "Income": {
            "total_label": "Total Income",
            "adjusted_total_label": "Adjusted Total Income",
            "omit_labels": ["Student Loans"],
            "header_padding": 1,
            "footer_padding": 1,
        },
        "Expenses": {
            "total_label": "Total Expenses",
            "adjusted_total_label": "Adjusted Total Expenses",
            "omit_labels": ["Investments", "Financing"],
            "header_padding": 1,
            "footer_padding": 1,
        },
        "Assets": {
            "total_label": "Net Assets",
            "header_padding": 1,
            "footer_padding": 1,
        },
        "Liabilities": {
            "total_label": "Net Liabilities",
            "header_padding": 1,
            "footer_padding": 1,
        },
    }

    root = ModelRow()

    for section in ["Income", "Expenses", "Assets", "Liabilities"]:
        categories = data[section]
        section_children = {}
        for category, entries in categories.items():
            category_children = {}
            for entry, entry_data in entries.items():
                entry_node = ModelRow(
                    label=entry,
                    data=entry_data,
                    style=_section_style_defaults[section]["Body"],
                )
                category_children[entry] = entry_node
            category_node = Total(
                label=category,
                children=category_children,
                style=_section_style_defaults[section]["Subtotal"],
            )
            section_children[category] = category_node

        section_node = Section(label=section, children=section_children)
        section_header = {
            section: ModelRow(
                label=section, style=_section_style_defaults[section]["Header"]
            )
        }
        if section in {"Income", "Expenses"}:
            section_footer = {
                "Total": Total(
                    label="Total",
                    style=_section_style_defaults[section]["Total"],
                    use_children_from=section_node,
                )
            }
        else:
            section_footer = {
                "Total": Total(
                    label="Total",
                    style=_section_style_defaults[section]["Total"],
                    use_children_from=section_node,
                ),
                "Adjusted Total": Total(
                    label="Adjusted Total",
                    style=_section_style_defaults[section]["Adjusted Total"],
                    use_children_from=section_node,
                ),
            }
        section_node.header = section_header
        section_node.footer = section_footer

        if section in {"Income", "Expenses"}:
            section_node = AdjustedTotalSection(
                label=section,
                children=section_children,
                style=_section_style_defaults[section]["Header"],
                total_style=_section_style_defaults[section]["Total"],
                adjusted_total_style=_section_style_defaults[section]["Adjusted Total"],
                **_section_kwarg_defaults[section]
            )
        else:
            section_node = TotalSection(
                label=section,
                children=section_children,
                style=_section_style_defaults[section]["Header"],
                total_style=_section_style_defaults[section]["Total"],
                **_section_kwarg_defaults[section]
            )
        root.children[section] = section_node
    return root


def main():
    import json

    with open("data.json") as f:
        data = json.load(f)
        root = build_model_tree(data)

    model = Model(root=root)
    model.set_zoom(160)
    model.save("model.xlsx")


if __name__ == "__main__":
    main()
