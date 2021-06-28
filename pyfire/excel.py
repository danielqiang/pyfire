import warnings
from typing import Union

import openpyxl
from openpyxl.styles import Alignment, Fill, Font
from openpyxl.utils import get_column_letter

ACCOUNTING_NUMBER_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


class Style:
    def __init__(
        self,
        font: Font = None,
        fill: Fill = None,
        alignment: Alignment = None,
        number_format: str = None,
    ):
        self.font = font
        self.fill = fill
        self.alignment = alignment
        self.number_format = number_format

    def apply(self, cell):
        if self.font:
            cell.font = self.font
        if self.fill:
            cell.fill = self.fill
        if self.alignment:
            cell.alignment = self.alignment
        if self.number_format:
            cell.number_format = self.number_format

    def apply_all(self, cells: list):
        for cell in cells:
            self.apply(cell)

    def __or__(self, other):
        if not isinstance(other, Style):
            return NotImplemented
        return Style(
            font=other.font or self.font,
            fill=other.fill or self.fill,
            alignment=other.alignment or self.alignment,
            number_format=other.number_format or self.number_format,
        )


class RowStyle:
    def __init__(
        self,
        label_style: Style = Style(),
        data_style: Style = Style(),
        row_total_style: Style = Style(),
    ):
        self.label_style = label_style
        self.data_style = data_style
        self.row_total_style = row_total_style

    def apply(self, row: list):
        if len(row) == 0:
            return

        self.label_style.apply(row[0])
        self.data_style.apply_all(row[1:-1])
        self.row_total_style.apply(row[-1])

    def __or__(self, other):
        if not isinstance(other, RowStyle):
            return NotImplemented
        return RowStyle(
            label_style=other.label_style | self.label_style,
            data_style=other.data_style | self.data_style,
            row_total_style=other.row_total_style | self.row_total_style,
        )


class ModelRow:
    HIDDEN = False

    # Default styling
    DEFAULT_LABEL_FONT: Font = None
    DEFAULT_LABEL_FILL: Fill = None
    DEFAULT_LABEL_ALIGNMENT: Alignment = None
    DEFAULT_LABEL_NUMBER_FORMAT: str = None

    DEFAULT_DATA_FONT: Font = None
    DEFAULT_DATA_FILL: Fill = None
    DEFAULT_DATA_ALIGNMENT: Alignment = None
    DEFAULT_DATA_NUMBER_FORMAT: str = ACCOUNTING_NUMBER_FORMAT

    DEFAULT_ROW_TOTAL_FONT: Font = None
    DEFAULT_ROW_TOTAL_FILL: Fill = None
    DEFAULT_ROW_TOTAL_ALIGNMENT: Alignment = None
    DEFAULT_ROW_TOTAL_NUMBER_FORMAT: str = ACCOUNTING_NUMBER_FORMAT

    def __init__(
        self,
        label: str = None,
        data: list[Union[int, str]] = None,
        children: dict[str, "ModelRow"] = None,
        style: RowStyle = RowStyle(),
        hidden: bool = False
    ):
        self.label = label
        self._data = data or []
        self.children = children or {}
        self.style = self.default_style() | style
        self.hidden = hidden or self.HIDDEN

        # Row is set by the model after the full model tree is finalized
        self._row = None

    def default_style(self) -> RowStyle:
        return RowStyle(
            label_style=Style(
                font=self.DEFAULT_LABEL_FONT,
                fill=self.DEFAULT_LABEL_FILL,
                alignment=self.DEFAULT_LABEL_ALIGNMENT,
                number_format=self.DEFAULT_LABEL_NUMBER_FORMAT,
            ),
            data_style=Style(
                font=self.DEFAULT_DATA_FONT,
                fill=self.DEFAULT_DATA_FILL,
                alignment=self.DEFAULT_DATA_ALIGNMENT,
                number_format=self.DEFAULT_DATA_NUMBER_FORMAT,
            ),
            row_total_style=Style(
                font=self.DEFAULT_ROW_TOTAL_FONT,
                fill=self.DEFAULT_ROW_TOTAL_FILL,
                alignment=self.DEFAULT_ROW_TOTAL_ALIGNMENT,
                number_format=self.DEFAULT_ROW_TOTAL_NUMBER_FORMAT,
            ),
        )

    @property
    def data(self):
        return self._data

    @data.setter
    def data(self, value):
        self._data = value

    @property
    def row(self):
        if self._row is None:
            raise AttributeError("Model row has not been set yet.")
        return self._row

    @row.setter
    def row(self, value):
        self._row = value

    @property
    def row_total(self):
        if len(self.data) == 0:
            formula = "0"
        else:
            start = get_column_letter(1)
            end = get_column_letter(len(self.data) + 1)
            formula = f"=SUM({start}{self.row},{end}{self.row})"
        return formula

    @property
    def height(self):
        if self.children:
            return 1 + max(child.height for child in self.children.values())
        else:
            return 0

    def iterator(self):
        if not self.hidden:
            yield self
        for child in self.children.values():
            yield from child.iterator()

    def __iter__(self):
        yield from self.iterator()


class Total(ModelRow):
    DEFAULT_LABEL_FONT = Font(bold=True)
    DEFAULT_DATA_FONT = Font(bold=True)
    DEFAULT_ROW_TOTAL_FONT = Font(bold=True)

    def __init__(
        self,
        *args,
        use_children_from: ModelRow = None,
        omit_labels: set[str] = None,
        **kwargs,
    ):
        super().__init__(*args, **kwargs)

        if use_children_from and self.children:
            warnings.warn("Children are ignored if 'use_children_from' is provided.")
        self.use_children_from = use_children_from
        self.omit_labels = omit_labels or set()

    @property
    def data(self):
        children = [
            child
            for child in (self.use_children_from or self).children.values()
            if child.label not in self.omit_labels
        ]

        if len(children) == 0:
            raise ValueError("Totals must have child nodes")
        if len(set(len(child.data) for child in children)) > 1:
            raise ValueError("Children must have identical data lengths")

        data = []
        for i in range(len(children[0].data)):
            column = get_column_letter(i)
            addresses = (f"{column}{child.row}" for child in children)
            formula = f"=SUM({','.join(addresses)})"
            data.append(formula)
        return data


class Section(ModelRow):
    HIDDEN = True

    DEFAULT_LABEL_FONT = Font(bold=True)
    DEFAULT_LABEL_ALIGNMENT = Alignment(horizontal="center")

    def __init__(
        self,
        *args,
        header: dict[str, ModelRow] = None,
        footer: dict[str, ModelRow] = None,
        **kwargs,
    ):
        super().__init__(*args, **kwargs)
        self.header = header or {}
        self.footer = footer or {}

    def iterator(self):
        yield from self.header.values()
        yield from super().iterator()
        yield from self.footer.values()


class Model:
    def __init__(self, root: ModelRow = None):
        self.root = root
        self.root.hidden = True

        self.wb = openpyxl.Workbook()

    def autofit_column_width(self):
        for sheet in self.wb.worksheets:
            for column_cell in next(sheet.rows):
                sheet.column_dimensions[column_cell.column_letter].bestFit = True

    def set_zoom(self, zoom: int):
        for sheet in self.wb.worksheets:
            sheet.sheet_view.zoomScale = zoom

    def save(self, filename: str):
        pass